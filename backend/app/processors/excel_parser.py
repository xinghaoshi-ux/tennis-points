from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.player_repo import PlayerRepository
from app.repositories.tournament_repo import TournamentRepository


RESULT_TYPE_MAP = {
    "冠军": "champion",
    "亚军": "runner_up",
    "四强": "semifinal",
    "八强": "quarterfinal",
    "参赛": "participant",
}

IS_TRUE_VALUES = {"是", "yes", "true", "1"}


class ExcelParser:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.player_repo = PlayerRepository(db)
        self.tournament_repo = TournamentRepository(db)

    def _normalize_result_type(self, raw: str) -> str:
        raw = raw.strip()
        if raw in RESULT_TYPE_MAP:
            return RESULT_TYPE_MAP[raw]
        return raw

    def _is_true(self, val) -> bool:
        if not val:
            return False
        return str(val).strip().lower() in IS_TRUE_VALUES

    async def parse(self, file_path: str, tournament_id: int) -> list[dict]:
        tournament = await self.tournament_repo.get_by_id(tournament_id)
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        preview = []

        for idx, row in enumerate(rows, start=1):
            if not row or all(cell is None for cell in row):
                continue

            raw_type = str(row[0]).strip() if row[0] else "participant"
            result_type = self._normalize_result_type(raw_type)
            player1_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            player2_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
            is_cross_province = self._is_true(row[3]) if len(row) > 3 else False
            is_cross_border = self._is_true(row[4]) if len(row) > 4 else False

            player1_id = None
            player1_matched = False
            player2_id = None
            player2_matched = False
            row_status = "normal"
            error_message = None

            if player1_name:
                p1 = await self.player_repo.get_by_name(player1_name)
                if p1:
                    player1_id = p1.id
                    player1_matched = True
                else:
                    row_status = "warning"
                    error_message = f"选手'{player1_name}'未匹配"

            if player2_name:
                p2 = await self.player_repo.get_by_name(player2_name)
                if p2:
                    player2_id = p2.id
                    player2_matched = True
                else:
                    if error_message:
                        error_message += f"；选手'{player2_name}'未匹配"
                    else:
                        error_message = f"选手'{player2_name}'未匹配"
                    row_status = "warning"

            preview.append({
                "row_number": idx,
                "tournament_name": tournament.name if tournament else None,
                "level": tournament.level if tournament else None,
                "group_name": tournament.group_name if tournament else None,
                "result_type": result_type,
                "player1_name": player1_name,
                "player1_id": player1_id,
                "player1_matched": player1_matched,
                "player2_name": player2_name,
                "player2_id": player2_id,
                "player2_matched": player2_matched,
                "is_cross_province": is_cross_province,
                "is_cross_border": is_cross_border,
                "estimated_points": None,
                "row_status": row_status,
                "error_message": error_message,
            })

        wb.close()
        return preview

    async def parse_team(self, file_path: str, tournament_id: int) -> list[dict]:
        """Parse team event Excel. Columns: 名次, 队伍名称, 队员1, 队员2, ..., 跨省, 跨境"""
        tournament = await self.tournament_repo.get_by_id(tournament_id)
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        preview = []

        for idx, row in enumerate(rows, start=1):
            if not row or all(cell is None for cell in row):
                continue

            raw_type = str(row[0]).strip() if row[0] else "participant"
            result_type = self._normalize_result_type(raw_type)
            team_name = str(row[1]).strip() if len(row) > 1 and row[1] else None

            members = []
            member_ids = []
            all_matched = True
            error_message = None

            for col_idx in range(2, len(row)):
                cell_val = row[col_idx]
                if not cell_val:
                    continue
                cell_str = str(cell_val).strip()
                if cell_str.lower() in IS_TRUE_VALUES:
                    break
                if cell_str in ("跨省", "跨境", ""):
                    break
                members.append(cell_str)

            is_cross_province = False
            is_cross_border = False
            if len(row) >= 2:
                last_two = [str(row[-2]).strip().lower() if row[-2] else "", str(row[-1]).strip().lower() if row[-1] else ""]
                if last_two[0] in IS_TRUE_VALUES:
                    is_cross_province = True
                if last_two[1] in IS_TRUE_VALUES:
                    is_cross_border = True

            for name in members:
                p = await self.player_repo.get_by_name(name)
                if p:
                    member_ids.append({"name": name, "id": p.id, "matched": True})
                else:
                    member_ids.append({"name": name, "id": None, "matched": False})
                    all_matched = False
                    if error_message:
                        error_message += f"；选手'{name}'未匹配"
                    else:
                        error_message = f"选手'{name}'未匹配"

            preview.append({
                "row_number": idx,
                "tournament_name": tournament.name if tournament else None,
                "level": tournament.level if tournament else None,
                "result_type": result_type,
                "team_name": team_name,
                "members": member_ids,
                "member_count": len(members),
                "is_cross_province": is_cross_province,
                "is_cross_border": is_cross_border,
                "row_status": "normal" if all_matched else "warning",
                "error_message": error_message,
            })

        wb.close()
        return preview
