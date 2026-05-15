from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError
from app.models.entries_points import EntriesPoints
from app.models.event_result_player import EventResultPlayer
from app.repositories.player_repo import PlayerRepository
from app.schemas.player import PlayerCreate, PlayerUpdate


class PlayerService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = PlayerRepository(db)

    async def list_players(
        self, page: int = 1, page_size: int = 20,
        search: str | None = None, department: str | None = None,
    ):
        return await self.repo.list(page, page_size, search, department)

    async def create_player(self, data: PlayerCreate):
        player = await self.repo.create(**data.model_dump())
        await self.db.commit()
        return player

    async def get_player(self, player_id: int):
        player = await self.repo.get_by_id(player_id)
        if not player:
            raise NotFoundError(detail="选手不存在", code="PLAYER_NOT_FOUND")
        return player

    async def update_player(self, player_id: int, data: PlayerUpdate):
        player = await self.repo.get_by_id(player_id)
        if not player:
            raise NotFoundError(detail="选手不存在", code="PLAYER_NOT_FOUND")

        update_data = data.model_dump(exclude_unset=True)
        player = await self.repo.update(player, **update_data)
        await self.db.commit()
        return player

    async def delete_player(self, player_id: int):
        player = await self.repo.get_by_id(player_id)
        if not player:
            raise NotFoundError(detail="选手不存在", code="PLAYER_NOT_FOUND")

        await self.db.execute(
            delete(EntriesPoints).where(EntriesPoints.player_id == player_id)
        )
        await self.db.execute(
            delete(EventResultPlayer).where(EventResultPlayer.player_id == player_id)
        )
        await self.repo.delete(player)
        await self.db.commit()

    async def batch_import(self, file) -> dict:
        from datetime import date
        from openpyxl import load_workbook
        import io

        GENDER_MAP = {"男": "male", "女": "female", "male": "male", "female": "female"}

        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            name = str(row[0]).strip() if row[0] else None
            if not name:
                errors.append(f"第{idx}行：姓名为空")
                continue

            existing = await self.repo.get_by_name(name)
            if existing:
                # Update info if provided
                updates = {}
                if len(row) > 1 and row[1]:
                    gender = GENDER_MAP.get(str(row[1]).strip().lower(), None)
                    if gender:
                        updates["gender"] = gender
                if len(row) > 2 and row[2]:
                    try:
                        bd = row[2]
                        if isinstance(bd, str):
                            bd = date.fromisoformat(bd.strip())
                        elif hasattr(bd, 'date'):
                            bd = bd.date()
                        updates["birth_date"] = bd
                    except (ValueError, AttributeError):
                        pass
                if len(row) > 3 and row[3]:
                    updates["department"] = str(row[3]).strip()
                if updates:
                    await self.repo.update(existing, **updates)
                skipped += 1
                continue

            kwargs = {"full_name": name}
            if len(row) > 1 and row[1]:
                gender = GENDER_MAP.get(str(row[1]).strip().lower(), None)
                if gender:
                    kwargs["gender"] = gender
            if len(row) > 2 and row[2]:
                try:
                    bd = row[2]
                    if isinstance(bd, str):
                        bd = date.fromisoformat(bd.strip())
                    elif hasattr(bd, 'date'):
                        bd = bd.date()
                    kwargs["birth_date"] = bd
                except (ValueError, AttributeError):
                    pass
            if len(row) > 3 and row[3]:
                kwargs["department"] = str(row[3]).strip()

            await self.repo.create(**kwargs)
            created += 1

        wb.close()
        await self.db.commit()

        return {"created": created, "skipped": skipped, "errors": errors}
